from datetime import datetime
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def open_order_list_window(self):
    from view.order_list_window import OrderListWindow
    self.order_list = OrderListWindow()
    self.order_list.show()
    self.close()


def export_stats_to_excel(self, stats, popular_dishes):
    current_date = datetime.now().strftime("%d_%m_%Y")

    default_filename = f"Статистика_на_{current_date}.xlsx"
    file_path, _ = QFileDialog.getSaveFileName(
        self, "Сохранить отчет", default_filename, "Excel Files (*.xlsx)"
    )

    if not file_path:
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Статистика за {current_date}"

        header_font = Font(bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells('A1:C1')
        ws['A1'] = f"Статистика ресторана за {current_date}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_alignment

        ws['A3'] = "Выполненных заказов:"
        ws['A3'].font = header_font
        ws['B3'] = stats[0] or 0

        ws['A4'] = "Средний чек (руб.):"
        ws['A4'].font = header_font
        ws['B4'] = round(stats[1], 2) if stats[1] is not None else 0

        ws['A5'] = "Выручка за день (руб.):"
        ws['A5'].font = header_font
        ws['B5'] = round(stats[2], 2) if stats[2] is not None else 0

        ws['A7'] = "Популярные блюда"
        ws['A7'].font = Font(bold=True, size=13)

        headers = ["Блюдо", "Количество", "Выручка (руб.)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        if popular_dishes:
            for row_idx, dish in enumerate(popular_dishes, 9):
                ws.cell(row=row_idx, column=1,
                        value=dish.dish.dish_name).border = thin_border
                ws.cell(row=row_idx, column=2,
                        value=dish.amount).border = thin_border
                ws.cell(row=row_idx, column=3, value=round(
                    dish.sum, 2)).border = thin_border

        column_widths = {'A': 30, 'B': 15, 'C': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(file_path)

        QMessageBox.information(self, "Экспорт успешен",
                                f"Отчет успешно сохранен в:\n{file_path}")

    except Exception as e:
        QMessageBox.critical(self, "Ошибка при экспорте",
                             f"Произошла ошибка при сохранении файла:\n{str(e)}")
